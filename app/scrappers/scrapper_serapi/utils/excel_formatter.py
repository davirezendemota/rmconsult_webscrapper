import os
import pandas as pd
import unicodedata
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from constants import RENOMEAR_COLUNAS


class ExcelFormatter:

    @staticmethod
    def gerar_excel(empresas, termo_busca):

        todas_chaves = set()
        for emp in empresas:
            for k in emp.keys():
                todas_chaves.add(k)

        todas_chaves = list(todas_chaves)

        prioridade = [
            "title", "address", "phone", "website",
            "agendar_on_line", "rating", "reviews",
            "gps_coordinates", "open_state"
        ]

        restantes = [c for c in todas_chaves if c not in prioridade]
        ordem_final = prioridade + sorted(restantes)

        linhas = []
        for emp in empresas:
            linha = {}
            for chave in ordem_final:
                valor = emp.get(chave, "")
                if valor is None:
                    valor = ""
                elif isinstance(valor, dict):
                    valor = str(valor)
                elif isinstance(valor, list):
                    valor = ", ".join(str(v) for v in valor)
                linha[chave] = valor

            linhas.append(linha)

        df = pd.DataFrame(linhas, columns=ordem_final)

        df.rename(columns={
            chave: RENOMEAR_COLUNAS.get(chave, chave.replace("_", " ").title())
            for chave in ordem_final
        }, inplace=True)

        nome_limpo = unicodedata.normalize("NFKD", termo_busca).encode("ASCII", "ignore").decode()
        nome_limpo = nome_limpo.replace(" ", "_")

        filename = f"{nome_limpo}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S-%f')}.xlsx"
        filepath = f"resultados/{filename}"

        os.makedirs("resultados", exist_ok=True)
        df.to_excel(filepath, index=False)

        wb = load_workbook(filepath)
        ws = wb.active
        ws.freeze_panes = "A2"

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for col_idx in range(1, ws.max_column + 1):
            c = ws.cell(1, col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")

        thin = Border(
            left=Side(style="thin", color="C0C0C0"),
            right=Side(style="thin", color="C0C0C0"),
            top=Side(style="thin", color="C0C0C0"),
            bottom=Side(style="thin", color="C0C0C0")
        )

        for row in range(2, ws.max_row + 1):
            fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") if row % 2 == 0 else None
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if fill:
                    cell.fill = fill
                cell.border = thin
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        for col_cells in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        table = Table(
            displayName="TabelaEmpresas",
            ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        )
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)

        wb.save(filepath)

        return filepath
